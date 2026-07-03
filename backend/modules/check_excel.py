"""check_excel.py

Diagnostic tool for validating a supervisor-provided uncertainty Excel file
before writing the category-specific parsing logic in formula_manager.py.

Legacy .xls files (the format Charkha has been sending) don't expose live
formulas through most Python libraries (pandas/xlrd only see the last
cached computed value) — so this script first converts to .xlsx via
LibreOffice headless, then inspects both the formula and its computed
value for every non-empty cell on the MU (Measurement Uncertainty) sheet.

Usage:
    python check_excel.py path/to/supervisor_file.xls

What it checks:
    1. The file converts cleanly to .xlsx (catches corrupt/password-protected files early).
    2. Every non-empty cell on the MU sheet is inspected: is it a formula,
       a hardcoded number, or text? Text cells in a region that should be
       numeric are flagged, since those are exactly the cells that will
       break a float-parsing formula_manager.py function later.
    3. Prints a coordinate-indexed report so you can build the cell-map
       config formula_manager.py needs, without guessing coordinates by eye.

This does not know in advance which cells matter for which instrument
category — it's a generic first pass, meant to be run once per new
supervisor file, with the output used to hand-write that category's
section of formula_manager.py.
"""

import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


def convert_to_xlsx(source_path: Path, out_dir: Path) -> Path:
    """Convert a legacy .xls file to .xlsx using headless LibreOffice.

    Args:
        source_path: Path to the source .xls (or .xlsx, passed through) file.
        out_dir: Directory to write the converted .xlsx file into.

    Returns:
        Path to the converted .xlsx file.

    Raises:
        RuntimeError: If the LibreOffice conversion fails or produces no output.
    """
    if source_path.suffix.lower() == ".xlsx":
        return source_path

    result = subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(source_path)],
        capture_output=True,
        text=True,
    )
    converted = out_dir / (source_path.stem + ".xlsx")
    if not converted.exists():
        raise RuntimeError(
            f"LibreOffice conversion failed for {source_path}.\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}"
        )
    return converted


def is_probably_numeric_field(label: str) -> bool:
    """Heuristic: does a nearby label suggest this cell should hold a number?

    Args:
        label: Text found in a nearby cell, likely a column/row header.

    Returns:
        bool: True if the label suggests the associated cell should be numeric.
    """
    numeric_hints = [
        "uncertainty", "reading", "value", "error", "cmc", "load", "resolution",
        "accuracy", "deviation", "std", "coefficient", "factor", "u", "k",
    ]
    label_lower = label.lower()
    return any(hint in label_lower for hint in numeric_hints)


def inspect_sheet(ws, sheet_name: str) -> dict:
    """Inspect all non-empty cells on a worksheet, separating formulas,
    numeric literals, and text.

    Args:
        ws: An openpyxl worksheet object (loaded with data_only=False so
            formulas are visible rather than only their cached values).
        sheet_name: Name of the sheet, used only for the report header.

    Returns:
        dict: Summary with keys 'formulas', 'numeric_literals', 'text_cells',
            and 'flagged' (text cells sitting in what looks like a numeric
            context, based on nearby labels — these are the ones most
            likely to break a float() parse downstream).
    """
    formulas = []
    numeric_literals = []
    text_cells = []
    flagged = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append((cell.coordinate, cell.value))
            elif isinstance(cell.value, (int, float)):
                numeric_literals.append((cell.coordinate, cell.value))
            elif isinstance(cell.value, str):
                text_cells.append((cell.coordinate, cell.value))
                # Check the cell immediately to the left and above for a
                # label that suggests this text cell is sitting where a
                # number is expected (e.g. "TBA" or "---" placeholders).
                left = ws.cell(row=cell.row, column=cell.column - 1).value if cell.column > 1 else None
                above = ws.cell(row=cell.row - 1, column=cell.column).value if cell.row > 1 else None
                nearby_label = " ".join(str(x) for x in [left, above] if isinstance(x, str))
                if nearby_label and is_probably_numeric_field(nearby_label):
                    flagged.append((cell.coordinate, cell.value, nearby_label))

    return {
        "sheet_name": sheet_name,
        "formulas": formulas,
        "numeric_literals": numeric_literals,
        "text_cells": text_cells,
        "flagged": flagged,
    }


def print_report(report: dict):
    """Print a human-readable report for one sheet's inspection results.

    Args:
        report: The dict returned by inspect_sheet.
    """
    print(f"\n{'=' * 60}")
    print(f"SHEET: {report['sheet_name']}")
    print(f"{'=' * 60}")
    print(f"  Formulas found:        {len(report['formulas'])}")
    print(f"  Numeric literals:      {len(report['numeric_literals'])}")
    print(f"  Text cells:            {len(report['text_cells'])}")
    print(f"  FLAGGED (text where numeric expected): {len(report['flagged'])}")

    if report["flagged"]:
        print("\n  --- Flagged cells (likely to break a float parse) ---")
        for coord, value, label in report["flagged"]:
            print(f"    {coord}: {value!r}  (near label: {label!r})")

    if report["formulas"]:
        print("\n  --- Sample formulas (first 15) ---")
        for coord, formula in report["formulas"][:15]:
            print(f"    {coord}: {formula}")


def main():
    if len(sys.argv) != 2:
        print("Usage: python check_excel.py path/to/supervisor_file.xls")
        sys.exit(1)

    source_path = Path(sys.argv[1])
    if not source_path.exists():
        print(f"File not found: {source_path}")
        sys.exit(1)

    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = convert_to_xlsx(source_path, Path(tmp_dir))
        wb = load_workbook(xlsx_path, data_only=False)

        print(f"File: {source_path.name}")
        print(f"Sheets found: {wb.sheetnames}")

        # MU (Measurement Uncertainty) is the sheet name both known
        # supervisor files use for the actual calculation. Fall back to
        # inspecting every sheet if MU isn't present, since a new category's
        # file might use a different sheet name.
        sheets_to_check = ["MU"] if "MU" in wb.sheetnames else wb.sheetnames

        for sheet_name in sheets_to_check:
            report = inspect_sheet(wb[sheet_name], sheet_name)
            print_report(report)

        print(f"\n{'=' * 60}")
        print("Done. Use the flagged cells above to confirm with the supervisor")
        print("whether they're intentional placeholders or missing data before")
        print("writing this category's formula_manager.py parsing logic.")


if __name__ == "__main__":
    main()
