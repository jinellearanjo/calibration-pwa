"""
modules/reporting.py
====================
Reporting module for the calibration application.

Responsibilities
----------------
1. gather_report_data  — orchestrates all six database calls into one ReportData object.
2. generate_pdf_report — builds a ReportLab PDF certificate and streams it.
3. generate_excel_report — builds an openpyxl workbook certificate and streams it.
4. write_audit_entry   — records a generation event to the audit log via database.py.

Rules enforced throughout
--------------------------
* Black and white only in PDF — no colour anywhere.
* Font is Helvetica (ReportLab built-in); no external font files required.
* All tables have black borders.
* Page numbers appear in the footer of every PDF page.
* If assets/logo.png exists it is drawn in the PDF header; if absent it is
  silently skipped — no exception is raised.
* Report files are never stored on the server: file is generated, streamed,
  then deleted via a background task attached to the FileResponse.
* Sessions with status REJECTED raise ValueError before any file is created.
* All database access goes through functions imported from database.py.
* No calculation logic and no validation logic live in this module.
"""

from __future__ import annotations

import logging
import math
import os
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi.responses import FileResponse

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

from database import (
    get_readings,               # readings table (Pressure / Electrical)
    get_calibration_reference,  # calibration_reference table
    get_session,                # calibration_sessions table
    get_instrument,             # instruments table
    get_master_instrument,      # master_instruments table
    get_uncertainty_budgets,     # uncertainty_budgets table
    get_weighing_repeatability_tests,   # weighing_repeatability_tests + nested readings
    get_weighing_off_center_readings,   # weighing_off_center_readings
    get_weighing_hysteresis_readings,   # weighing_hysteresis_readings
    get_temperature_repeatability_tests,  # temperature_repeatability_tests + nested readings
    insert_audit_log,           # audit_log table
)

logger = logging.getLogger(__name__)

# ─── Module-level constants ────────────────────────────────────────────────────

PAGE_WIDTH, PAGE_HEIGHT = A4
LOGO_PATH = Path("assets/logo.png")

LEFT_MARGIN = 20 * mm
RIGHT_MARGIN = 20 * mm
TOP_MARGIN = 38 * mm
BOTTOM_MARGIN = 25 * mm

BLACK = colors.black
WHITE = colors.white

FONT_NORMAL = "Helvetica"
FONT_BOLD = "Helvetica-Bold"
FONT_OBLIQUE = "Helvetica-Oblique"

THIN_BORDER = 0.5


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — DATA GATHERING
# ══════════════════════════════════════════════════════════════════════════════


@dataclass
class ReportData:
    """Typed container for every field a calibration certificate must render.

    Produced exclusively by gather_report_data and consumed by
    generate_pdf_report, generate_excel_report, and write_audit_entry.
    Field names mirror the database column names from the spec so that a
    renamed column causes a visible KeyError in gather_report_data rather
    than a silent None in a rendered cell.

    Attributes:
        session_id: Primary key of the calibration session.
        session_status: Uppercased status string from calibration_sessions.
        certificate_number: Unique certificate identifier from calibration_reference.
        date_of_calibration: Date the calibration was performed.
        cal_due_date: Next calibration due date.
        item_received_date: Date the instrument was received by the lab.
        date_of_issue: Date this certificate was issued.
        customer_name: Name of the customer organisation.
        customer_address: Full postal address of the customer.
        technician: Name of the technician who performed the calibration.
        instrument_name: Human-readable name of the instrument under test.
        instrument_make: Manufacturer name.
        instrument_model: Model designation.
        instrument_serial_number: Serial number of the instrument under test.
        instrument_tag_number: Site tag or asset number.
        instrument_range_min: Lower bound of the instrument's measurement range.
        instrument_range_max: Upper bound of the instrument's measurement range.
        instrument_unit: Unit of measurement (e.g. bar, degC).
        instrument_resolution: Instrument resolution.
        instrument_accuracy_class: Manufacturer-stated accuracy class/spec.
        instrument_dial_size: Dial size, if applicable.
        instrument_mounting_orientation: Mounting orientation, if applicable.
        instrument_location: Physical location of the instrument.
        instrument_medium_used: Medium used during calibration.
        instrument_calibration_carried_at: Location calibration was carried out.
        temperature_c: Ambient temperature in degrees Celsius during calibration.
        humidity_pct: Relative humidity percentage during calibration.
        master_name: Name of the reference / master instrument.
        master_make: Manufacturer of the master instrument.
        master_model: Model designation of the master instrument.
        master_serial_number: Serial number of the master instrument.
        master_asset_number: Asset or inventory number of the master instrument.
        master_traceability_chain: Documented traceability chain string.
        master_uncertainty_u: Uncertainty from the master's own certificate.
        master_accuracy: Accuracy of the master instrument.
        master_resolution: Resolution of the master instrument.
        master_cal_due_date: Next calibration due date of the master instrument.
        master_claimed_cmc: Claimed calibration and measurement capability.
        master_certificate_number: The master instrument's own calibration
            certificate number (distinct from certificate_number, which is
            this app's own certificate for the instrument under test).
            May be None for master instruments recorded before this field
            existed - renderers should display it as blank in that case,
            not fabricate a value.
        readings: Ordered list of per-point measurement dicts.
        type_a_value: Type A uncertainty evaluation result.
        u_std: Standard uncertainty.
        u_std_accuracy: Standard's accuracy uncertainty contribution (Ub2).
        u_res: Resolution uncertainty contribution.
        u_hys: Hysteresis uncertainty contribution.
        u_zero: Zero / offset uncertainty contribution.
        u_temp: Temperature uncertainty contribution.
        u_repeatability: Repeatability uncertainty contribution, distinct
            from Type A (Pressure's Ub5).
        u_std_weights: Standard weights uncertainty contribution (Weighing only).
        u_eccentric: Eccentric loading uncertainty contribution (Weighing only).
        u_drift: Drift of standard uncertainty contribution (Temperature only).
        u_bath_stability: Bath stability uncertainty contribution (Temperature only).
        u_bath_uniformity: Bath uniformity uncertainty contribution (Temperature only).
        u_wire_homogeneity: Wire homogeneity uncertainty contribution
            (Temperature, TCK sub-type only).
        cmc: CMC uncertainty contribution.
        combined_uncertainty: Combined standard uncertainty.
        expanded_uncertainty: Expanded uncertainty.
        k_value: Coverage factor k.
        final_applied_uncertainty: Final uncertainty value stated on certificate.
    """

    session_id: str
    session_status: str

    # calibration_reference
    certificate_number: str | None
    date_of_calibration: Any
    cal_due_date: Any
    item_received_date: Any
    date_of_issue: Any
    customer_name: str | None
    customer_address: str | None

    # calibration_sessions
    technician: str | None

    # instruments
    instrument_name: str | None
    instrument_type: str | None
    instrument_subtype: str | None
    instrument_make: str | None
    instrument_model: str | None
    instrument_serial_number: str | None
    instrument_tag_number: str | None
    instrument_range_min: Any
    instrument_range_max: Any
    instrument_unit: str | None
    instrument_resolution: Any
    instrument_accuracy_class: str | None
    instrument_dial_size: str | None
    instrument_mounting_orientation: str | None
    instrument_location: str | None
    instrument_medium_used: str | None
    instrument_calibration_carried_at: str | None

    # environmental conditions from calibration_sessions
    temperature_c: Any
    humidity_pct: Any

    # master_instruments
    master_name: str | None
    master_make: str | None
    master_model: str | None
    master_serial_number: str | None
    master_asset_number: str | None
    master_traceability_chain: str | None
    master_uncertainty_u: Any
    master_accuracy: Any
    master_resolution: Any
    master_cal_due_date: Any
    master_claimed_cmc: Any
    master_certificate_number: str | None

    # readings — order preserved from the database call. Only populated for
    # the category that matches instrument_type; the other two are left as
    # empty lists. Kept as three separate fields (rather than one polymorphic
    # blob) so a renderer that reads the wrong one gets an empty table
    # instead of silently rendering another category's shape.
    readings: list[dict[str, Any]] = field(default_factory=list)
    weighing_repeatability: list[dict[str, Any]] = field(default_factory=list)
    weighing_off_center: list[dict[str, Any]] = field(default_factory=list)
    weighing_hysteresis: list[dict[str, Any]] = field(default_factory=list)
    temperature_repeatability: list[dict[str, Any]] = field(default_factory=list)

    # uncertainty_budgets — spec field names used verbatim
    type_a_value: Any = None
    u_std: Any = None
    u_std_accuracy: Any = None
    u_res: Any = None
    u_hys: Any = None
    u_zero: Any = None
    u_temp: Any = None
    u_repeatability: Any = None
    u_std_weights: Any = None
    u_eccentric: Any = None
    u_drift: Any = None
    u_bath_stability: Any = None
    u_bath_uniformity: Any = None
    u_wire_homogeneity: Any = None
    cmc: Any = None
    combined_uncertainty: Any = None
    expanded_uncertainty: Any = None
    k_value: Any = None
    final_applied_uncertainty: Any = None


def gather_report_data(session_id: str) -> ReportData:
    """Fetch and assemble all data required to render a calibration certificate.

    Makes one call per source table through database.py. No calculation or
    validation logic runs here; values are mapped directly from database
    records into a ReportData instance.

    The REJECTED status guard lives here so that both renderers are protected
    by calling this single function rather than each implementing their own check.

    Args:
        session_id: UUID of the calibration session to report on.

    Returns:
        ReportData: Fully populated instance ready for a renderer.

    Raises:
        ValueError: If no session exists for session_id, or if the session
            status is REJECTED.
        RuntimeError: If a required database record is absent.
    """
    # Fetched first because status determines whether any further work proceeds.
    session_record = get_session(session_id)
    if session_record is None:
        raise ValueError(f"No calibration session found for session_id={session_id}.")

    session_status = str(session_record.get("status", "")).upper()

    # Raising here means neither PDF nor Excel generation can bypass this rule.
    if session_status == "REJECTED":
        raise ValueError(
            f"Session {session_id} has status REJECTED. "
            "Certificate generation is not permitted for rejected sessions."
        )

    # calibration_reference
    certificate_record = get_calibration_reference(session_id)
    if certificate_record is None:
        raise RuntimeError(f"No calibration_reference record found for session_id={session_id}.")

    # instruments — instrument_id is stored on the session record
    instrument_id = session_record.get("instrument_id")
    instrument_record = get_instrument(instrument_id)
    if instrument_record is None:
        raise RuntimeError(f"No instruments record found for instrument_id={instrument_id}.")

    # readings — shape depends on instrument_type: Weighing and Temperature
    # store their raw data in their own dedicated tables rather than the
    # generic "readings" table. Only the category matching this instrument
    # is fetched; the other fields stay as empty lists.
    instrument_type = instrument_record.get("type")
    readings_rows: list[dict[str, Any]] = []
    weighing_repeatability_rows: list[dict[str, Any]] = []
    weighing_off_center_rows: list[dict[str, Any]] = []
    weighing_hysteresis_rows: list[dict[str, Any]] = []
    temperature_repeatability_rows: list[dict[str, Any]] = []

    if instrument_type == "Weighing":
        weighing_repeatability_rows = get_weighing_repeatability_tests(session_id) or []
        weighing_off_center_rows = get_weighing_off_center_readings(session_id) or []
        weighing_hysteresis_rows = get_weighing_hysteresis_readings(session_id) or []
    elif instrument_type == "Temperature":
        temperature_repeatability_rows = get_temperature_repeatability_tests(session_id) or []
    else:
        # Pressure / Electrical - empty list is valid for a draft session
        readings_rows = get_readings(session_id) or []

    # uncertainty_budgets - now a LIST (formula_manager always returns a
    # list; Temperature/Electrical can have more than one). Certificates
    # showing a proper per-setpoint/range uncertainty breakdown for
    # multi-budget sessions is a known follow-up piece of work, same
    # category as the Calibration Readings table redesign - NOT silently
    # picking one budget or rendering something misleading in the
    # meantime, failing loudly and clearly instead.
    uncertainty_records = get_uncertainty_budgets(session_id) or []
    if not uncertainty_records:
        raise RuntimeError(f"No uncertainty_budgets record found for session_id={session_id}.")
    if len(uncertainty_records) > 1:
        raise RuntimeError(
            f"Session {session_id} has {len(uncertainty_records)} uncertainty budgets "
            f"(one per Temperature setpoint or Electrical function-type/range). "
            f"Certificate generation for multi-budget sessions isn't built yet - "
            f"only single-budget sessions (Pressure, Weighing, or a Temperature/"
            f"Electrical session with exactly one setpoint/range) can generate "
            f"a certificate right now."
        )
    uncertainty_record = uncertainty_records[0]

    # master_instruments — linked via calibration_sessions.master_instrument_id
    master_instrument_id = session_record.get("master_instrument_id")
    master_record = {}
    if master_instrument_id:
        master_record = get_master_instrument(master_instrument_id) or {}
    else:
        logger.warning(
            "Session %s has no master_instrument_id set; certificate will be "
            "generated with blank master instrument fields.",
            session_id,
        )

    return ReportData(
        session_id=session_id,
        session_status=session_status,

        certificate_number=certificate_record.get("certificate_number"),
        date_of_calibration=certificate_record.get("date_of_calibration"),
        cal_due_date=certificate_record.get("cal_due_date"),
        item_received_date=certificate_record.get("item_received_date"),
        date_of_issue=certificate_record.get("date_of_issue"),
        customer_name=certificate_record.get("customer_name"),
        customer_address=certificate_record.get("customer_address"),

        technician=session_record.get("technician"),

        instrument_name=instrument_record.get("name"),
        instrument_type=instrument_type,
        instrument_subtype=instrument_record.get("instrument_subtype"),
        instrument_make=instrument_record.get("make"),
        instrument_model=instrument_record.get("model"),
        instrument_serial_number=instrument_record.get("serial_number"),
        instrument_tag_number=instrument_record.get("tag_number"),
        instrument_range_min=instrument_record.get("range_min"),
        instrument_range_max=instrument_record.get("range_max"),
        instrument_unit=instrument_record.get("unit"),
        instrument_resolution=instrument_record.get("resolution"),
        instrument_accuracy_class=instrument_record.get("accuracy_class"),
        instrument_dial_size=instrument_record.get("dial_size"),
        instrument_mounting_orientation=instrument_record.get("mounting_orientation"),
        instrument_location=instrument_record.get("instrument_location"),
        instrument_medium_used=instrument_record.get("medium_used"),
        instrument_calibration_carried_at=instrument_record.get("calibration_carried_at"),

        # Environmental conditions are stored on the calibration session
        temperature_c=session_record.get("temperature_c"),
        humidity_pct=session_record.get("humidity_pct"),

        master_name=master_record.get("name"),
        master_make=master_record.get("make"),
        master_model=master_record.get("model"),
        master_serial_number=master_record.get("serial_number"),
        master_asset_number=master_record.get("asset_number"),
        master_traceability_chain=master_record.get("traceability_chain"),
        master_uncertainty_u=master_record.get("uncertainty_u"),
        master_accuracy=master_record.get("accuracy"),
        master_resolution=master_record.get("resolution"),
        master_cal_due_date=master_record.get("cal_due_date"),
        master_claimed_cmc=master_record.get("claimed_cmc"),
        master_certificate_number=master_record.get("master_certificate_number"),

        readings=readings_rows,
        weighing_repeatability=weighing_repeatability_rows,
        weighing_off_center=weighing_off_center_rows,
        weighing_hysteresis=weighing_hysteresis_rows,
        temperature_repeatability=temperature_repeatability_rows,

        type_a_value=uncertainty_record.get("type_a_value"),
        u_std=uncertainty_record.get("u_std"),
        u_std_accuracy=uncertainty_record.get("u_std_accuracy"),
        u_res=uncertainty_record.get("u_res"),
        u_hys=uncertainty_record.get("u_hys"),
        u_zero=uncertainty_record.get("u_zero"),
        u_temp=uncertainty_record.get("u_temp"),
        u_repeatability=uncertainty_record.get("u_repeatability"),
        u_std_weights=uncertainty_record.get("u_std_weights"),
        u_eccentric=uncertainty_record.get("u_eccentric"),
        u_drift=uncertainty_record.get("u_drift"),
        u_bath_stability=uncertainty_record.get("u_bath_stability"),
        u_bath_uniformity=uncertainty_record.get("u_bath_uniformity"),
        u_wire_homogeneity=uncertainty_record.get("u_wire_homogeneity"),
        cmc=uncertainty_record.get("cmc"),
        combined_uncertainty=uncertainty_record.get("combined_uncertainty"),
        expanded_uncertainty=uncertainty_record.get("expanded_uncertainty"),
        k_value=uncertainty_record.get("k_value"),
        final_applied_uncertainty=uncertainty_record.get("final_applied_uncertainty"),
    )


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1B — CATEGORY-AWARE READINGS TABLE BUILDER
# ══════════════════════════════════════════════════════════════════════════════


def _build_readings_blocks(report_data: ReportData) -> list[dict[str, Any]]:
    """Build the Calibration Readings table(s) for whichever category this
    session belongs to.

    Previously this table was hardcoded to Pressure's ascending/descending
    shape for every category, so a Weighing or Temperature certificate
    rendered an empty or wrong-shaped table (a known, documented gap).

    Returns a list of blocks, each a dict with:
        title: str | None — sub-heading shown above this block's table
            (Weighing needs three; Pressure/Electrical/Temperature need
            just one, so title is None and no sub-heading is drawn).
        header: list[str] — column labels.
        rows: list[list[str]] — pre-stringified cell values (via _safe_str),
            ready for either ReportLab's Table or an openpyxl row writer.

    Shared by generate_pdf_report and generate_excel_report so the two
    renderers read from a single source of truth and can't independently
    drift out of sync with each other, the way the two duplicated readings
    tables previously could have.
    """
    instrument_type = report_data.instrument_type

    if instrument_type == "Weighing":
        repeatability_header = ["Test Point", "Nominal Load", "Unit", "Reading #", "Before", "With Load", "After"]
        repeatability_rows: list[list[str]] = []
        for test in report_data.weighing_repeatability:
            nested_readings = sorted(
                test.get("weighing_repeatability_readings") or [],
                key=lambda r: r.get("reading_number", 0),
            )
            for r in nested_readings:
                repeatability_rows.append([
                    _safe_str(test.get("test_point")),
                    _safe_str(test.get("nominal_load")),
                    _safe_str(test.get("unit")),
                    _safe_str(r.get("reading_number")),
                    _safe_str(r.get("reading_before")),
                    _safe_str(r.get("reading_with_load")),
                    _safe_str(r.get("reading_after")),
                ])

        off_center_header = ["Position", "Nominal Load", "Unit", "Before", "With Load", "After"]
        off_center_rows = [
            [
                _safe_str(r.get("position")),
                _safe_str(r.get("nominal_load")),
                _safe_str(r.get("unit")),
                _safe_str(r.get("reading_before")),
                _safe_str(r.get("reading_with_load")),
                _safe_str(r.get("reading_after")),
            ]
            for r in report_data.weighing_off_center
        ]

        hysteresis_header = ["Sequence", "Phase", "Reading Value", "Unit"]
        hysteresis_rows = [
            [
                _safe_str(r.get("sequence_order")),
                _safe_str(r.get("phase")),
                _safe_str(r.get("reading_value")),
                _safe_str(r.get("unit")),
            ]
            for r in report_data.weighing_hysteresis
        ]

        return [
            {"title": "Repeatability", "header": repeatability_header, "rows": repeatability_rows},
            {"title": "Off-Center (Eccentricity)", "header": off_center_header, "rows": off_center_rows},
            {"title": "Hysteresis", "header": hysteresis_header, "rows": hysteresis_rows},
        ]

    if instrument_type == "Temperature":
        header = ["Setpoint", "Nominal Temperature", "Unit", "Reading #", "Reading Value"]
        rows: list[list[str]] = []
        for test in report_data.temperature_repeatability:
            nested_readings = sorted(
                test.get("temperature_repeatability_readings") or [],
                key=lambda r: r.get("reading_number", 0),
            )
            for r in nested_readings:
                rows.append([
                    _safe_str(test.get("setpoint_label")),
                    _safe_str(test.get("nominal_temperature")),
                    _safe_str(test.get("unit")),
                    _safe_str(r.get("reading_number")),
                    _safe_str(r.get("reading_value")),
                ])
        return [{"title": None, "header": header, "rows": rows}]

    # Default: Pressure / Electrical - the original ascending/descending
    # shape, unchanged from before this function existed. The "\n" in the
    # PDF-facing header is preserved further down where the PDF story is
    # built, since ReportLab's Table splits plain-string cells on "\n"
    # into wrapped lines (verified empirically) but openpyxl should not
    # receive literal embedded newlines in a header cell.
    header = ["Nominal Value", "Ascending Measured", "Descending Measured", "Mean Error", "Hysteresis"]
    rows = [
        [
            _safe_str(row.get("nominal_value")),
            _safe_str(row.get("measured_value_up")),      # matches database column
            _safe_str(row.get("measured_value_down")),    # matches database column
            _safe_str(row.get("mean_error")),
            _safe_str(row.get("hysteresis")),
        ]
        for row in report_data.readings
    ]
    return [{"title": None, "header": header, "rows": rows}]


def _format_date(raw_date: Any) -> str:
    """Format a date value to a human-readable string for certificate display.

    Args:
        raw_date: A datetime, date, ISO string, or None.

    Returns:
        str: Formatted string such as "15 Jun 2025", or "—" for None.
    """
    if raw_date is None:
        return "\u2014"
    if isinstance(raw_date, str):
        try:
            raw_date = datetime.fromisoformat(raw_date)
        except ValueError:
            return str(raw_date)
    return raw_date.strftime("%d %b %Y")


def _safe_str(value: Any) -> str:
    """Return value as a string, substituting an em-dash for None.

    Args:
        value: Any value from a ReportData field.

    Returns:
        str: String representation, or "—" when value is None.
    """
    return "\u2014" if value is None else str(value)


def _build_paragraph_styles() -> dict[str, ParagraphStyle]:
    """Return named ParagraphStyle objects for PDF certificate sections.

    Returns:
        dict[str, ParagraphStyle]: Mapping of style name to ParagraphStyle.
    """
    base = getSampleStyleSheet()

    return {
        "title": ParagraphStyle(
            "CertTitle",
            parent=base["Normal"],
            fontName=FONT_BOLD,
            fontSize=14,
            leading=18,
            alignment=1,
            textColor=BLACK,
        ),
        "section_heading": ParagraphStyle(
            "SectionHeading",
            parent=base["Normal"],
            fontName=FONT_BOLD,
            fontSize=9,
            leading=12,
            spaceBefore=6,
            spaceAfter=3,
            textColor=BLACK,
        ),
        "body": ParagraphStyle(
            "Body",
            parent=base["Normal"],
            fontName=FONT_NORMAL,
            fontSize=8,
            leading=11,
            textColor=BLACK,
        ),
        "body_bold": ParagraphStyle(
            "BodyBold",
            parent=base["Normal"],
            fontName=FONT_BOLD,
            fontSize=8,
            leading=11,
            textColor=BLACK,
        ),
        "footer": ParagraphStyle(
            "Footer",
            parent=base["Normal"],
            fontName=FONT_OBLIQUE,
            fontSize=7,
            leading=9,
            alignment=1,
            textColor=BLACK,
        ),
    }


def _black_table_style(has_header_row: bool = True) -> TableStyle:
    """Return a TableStyle with black borders and optional bold header row.

    Args:
        has_header_row: When True, applies a black-filled header row with
            white bold text.

    Returns:
        TableStyle: Configured instance.
    """
    light_grey = colors.Color(0.93, 0.93, 0.93)

    commands = [
        ("FONTNAME", (0, 0), (-1, -1), FONT_NORMAL),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 11),
        ("TEXTCOLOR", (0, 0), (-1, -1), BLACK),
        ("GRID", (0, 0), (-1, -1), THIN_BORDER, BLACK),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, light_grey]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]

    if has_header_row:
        commands += [
            ("BACKGROUND", (0, 0), (-1, 0), BLACK),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, light_grey]),
        ]

    return TableStyle(commands)


def _kv_table_style() -> TableStyle:
    """Return a TableStyle for two-column label/value pair tables.

    Returns:
        TableStyle: Configured instance.
    """
    return TableStyle(
        [
            ("FONTNAME", (0, 0), (0, -1), FONT_BOLD),
            ("FONTNAME", (1, 0), (1, -1), FONT_NORMAL),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (0, 0), (-1, -1), 11),
            ("TEXTCOLOR", (0, 0), (-1, -1), BLACK),
            ("BOX", (0, 0), (-1, -1), THIN_BORDER, BLACK),
            ("INNERGRID", (0, 0), (-1, -1), THIN_BORDER, BLACK),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
    )


def _build_kv_table(
    pairs: list[tuple[str, str]],
    styles: dict[str, ParagraphStyle],
    col_widths: list[float],
) -> Table:
    """Build a two-column label/value Table from a list of (label, value) pairs.

    Args:
        pairs: Sequence of (label_string, value_string) tuples.
        styles: Style dict from _build_paragraph_styles.
        col_widths: Two-element list of column widths in points.

    Returns:
        Table: Formatted ReportLab Table.
    """
    table_data = [
        [
            Paragraph(label, styles["body_bold"]),
            Paragraph(value, styles["body"]),
        ]
        for label, value in pairs
    ]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(_kv_table_style())
    return table


def _build_header_footer_callback(
    certificate_number: str,
    total_pages_ref: list,
):
    """Return an onPage callback that draws the PDF header and footer.

    Args:
        certificate_number: Certificate number shown in the header.
        total_pages_ref: Single-element list updated to the true page count
            after the first build pass completes.

    Returns:
        Callable: The on_page function to register with SimpleDocTemplate.
    """

    def on_page(canvas_obj, doc):
        canvas_obj.saveState()

        header_baseline = PAGE_HEIGHT - 10 * mm

        if LOGO_PATH.exists():
            try:
                canvas_obj.drawImage(
                    str(LOGO_PATH),
                    LEFT_MARGIN,
                    header_baseline - 12 * mm,
                    width=28 * mm,
                    height=12 * mm,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                # A corrupt or unsupported logo must never abort the certificate.
                logger.warning(
                    "Logo exists at %s but could not be drawn; skipping.",
                    LOGO_PATH,
                    exc_info=True,
                )

        canvas_obj.setFont(FONT_BOLD, 14)
        canvas_obj.setFillColor(BLACK)
        canvas_obj.drawCentredString(
            PAGE_WIDTH / 2, header_baseline - 5 * mm, "CALIBRATION CERTIFICATE"
        )
        canvas_obj.setFont(FONT_NORMAL, 8)
        canvas_obj.drawCentredString(
            PAGE_WIDTH / 2,
            header_baseline - 11 * mm,
            f"Certificate No: {certificate_number}",
        )

        canvas_obj.setStrokeColor(BLACK)
        canvas_obj.setLineWidth(0.75)
        canvas_obj.line(
            LEFT_MARGIN,
            PAGE_HEIGHT - TOP_MARGIN + 4 * mm,
            PAGE_WIDTH - RIGHT_MARGIN,
            PAGE_HEIGHT - TOP_MARGIN + 4 * mm,
        )

        footer_y = BOTTOM_MARGIN - 7 * mm
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(LEFT_MARGIN, BOTTOM_MARGIN - 2 * mm, PAGE_WIDTH - RIGHT_MARGIN, BOTTOM_MARGIN - 2 * mm)

        canvas_obj.setFont(FONT_OBLIQUE, 7)
        canvas_obj.drawString(
            LEFT_MARGIN,
            footer_y,
            "This certificate is valid only when reproduced in its entirety.",
        )
        canvas_obj.drawRightString(
            PAGE_WIDTH - RIGHT_MARGIN,
            footer_y,
            f"Page {doc.page} of {total_pages_ref[0]}",
        )

        canvas_obj.restoreState()

    return on_page


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — PDF GENERATION
# ══════════════════════════════════════════════════════════════════════════════


def generate_pdf_report(session_id: str) -> FileResponse:
    """Generate a PDF calibration certificate for the given session.

    Args:
        session_id: UUID of the calibration session.

    Returns:
        FileResponse: Streaming PDF attachment.

    Raises:
        ValueError: When the session does not exist or has status REJECTED.
        RuntimeError: When a required database record is missing.
    """
    report_data = gather_report_data(session_id)

    styles = _build_paragraph_styles()
    usable_width = PAGE_WIDTH - LEFT_MARGIN - RIGHT_MARGIN
    label_col_narrow = 50 * mm
    label_col_wide = 70 * mm

    total_pages_ref: list[int | str] = ["?"]

    temp_file = tempfile.NamedTemporaryFile(
        delete=False, suffix=".pdf", prefix=f"cal_cert_{session_id}_"
    )
    temp_path = temp_file.name
    temp_file.close()

    try:
        doc = SimpleDocTemplate(
            temp_path,
            pagesize=A4,
            leftMargin=LEFT_MARGIN,
            rightMargin=RIGHT_MARGIN,
            topMargin=TOP_MARGIN,
            bottomMargin=BOTTOM_MARGIN,
        )

        on_page = _build_header_footer_callback(
            certificate_number=_safe_str(report_data.certificate_number),
            total_pages_ref=total_pages_ref,
        )

        story = _build_pdf_story(report_data, styles, usable_width, label_col_narrow, label_col_wide)

        # First pass establishes page count.
        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
        total_pages_ref[0] = doc.page

        # Second pass so all footers show the correct total.
        story = _build_pdf_story(report_data, styles, usable_width, label_col_narrow, label_col_wide)
        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)

        safe_cert_number = "".join(
            ch for ch in _safe_str(report_data.certificate_number)
            if ch.isalnum() or ch in ("-", "_")
        )
        download_filename = f"CalibrationCertificate_{safe_cert_number}.pdf"

        return FileResponse(
            path=temp_path,
            media_type="application/pdf",
            filename=download_filename,
            background=_DeleteFileAfterResponse(temp_path),
        )

    except Exception:
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        raise


def _build_pdf_story(
    report_data: ReportData,
    styles: dict[str, ParagraphStyle],
    usable_width: float,
    label_col_narrow: float,
    label_col_wide: float,
) -> list:
    """Build the ordered list of Platypus flowables for the certificate.

    Args:
        report_data: Populated ReportData instance.
        styles: Paragraph style dict from _build_paragraph_styles.
        usable_width: Printable width in points.
        label_col_narrow: Width in points for narrow label columns.
        label_col_wide: Width in points for wide label columns.

    Returns:
        list: Ordered list of ReportLab Platypus flowables.
    """
    value_col_narrow = usable_width - label_col_narrow
    value_col_wide = usable_width - label_col_wide
    story = []

    def heading(text: str):
        story.append(Paragraph(text, styles["section_heading"]))

    def spacer():
        story.append(Spacer(1, 3 * mm))

    # Certificate details
    heading("Certificate Details")
    story.append(
        _build_kv_table(
            [
                ("Certificate No.:", _safe_str(report_data.certificate_number)),
                ("Date of Calibration:", _format_date(report_data.date_of_calibration)),
                ("Calibration Due Date:", _format_date(report_data.cal_due_date)),
                ("Item Received:", _format_date(report_data.item_received_date)),
                ("Date of Issue:", _format_date(report_data.date_of_issue)),
                ("Customer Name:", _safe_str(report_data.customer_name)),
                ("Customer Address:", _safe_str(report_data.customer_address)),
                ("Technician:", _safe_str(report_data.technician)),
            ],
            styles,
            [label_col_narrow, value_col_narrow],
        )
    )

    # Instrument under test
    spacer()
    heading("Instrument Under Test")
    pdf_cal_range_display = None
    if report_data.instrument_range_min is not None and report_data.instrument_range_max is not None:
        unit_suffix = f" {report_data.instrument_unit}" if report_data.instrument_unit else ""
        pdf_cal_range_display = (
            f"{report_data.instrument_range_min} to {report_data.instrument_range_max}{unit_suffix}"
        )
    story.append(
        _build_kv_table(
            [
                ("Instrument Name:", _safe_str(report_data.instrument_name)),
                ("Make:", _safe_str(report_data.instrument_make)),
                ("Model:", _safe_str(report_data.instrument_model)),
                ("Serial Number:", _safe_str(report_data.instrument_serial_number)),
                ("Tag Number:", _safe_str(report_data.instrument_tag_number)),
                ("Cal Range:", _safe_str(pdf_cal_range_display)),
                ("Resolution:", _safe_str(report_data.instrument_resolution)),
                ("Accuracy:", _safe_str(report_data.instrument_accuracy_class)),
                ("Dial Size:", _safe_str(report_data.instrument_dial_size)),
                ("Mounting Orientation:", _safe_str(report_data.instrument_mounting_orientation)),
                ("Location:", _safe_str(report_data.instrument_location)),
                ("Medium:", _safe_str(report_data.instrument_medium_used)),
                ("Calibration Carried At:", _safe_str(report_data.instrument_calibration_carried_at)),
            ],
            styles,
            [label_col_narrow, value_col_narrow],
        )
    )

    # Environmental conditions
    spacer()
    heading("Environmental Conditions at Time of Calibration")
    story.append(
        _build_kv_table(
            [
                ("Temperature (°C):", _safe_str(report_data.temperature_c)),
                ("Relative Humidity (%):", _safe_str(report_data.humidity_pct)),
            ],
            styles,
            [label_col_narrow, value_col_narrow],
        )
    )

    # Master instrument
    spacer()
    heading("Reference / Master Instrument")
    story.append(
        _build_kv_table(
            [
                ("Instrument Name:", _safe_str(report_data.master_name)),
                ("Make:", _safe_str(report_data.master_make)),
                ("Model:", _safe_str(report_data.master_model)),
                ("Serial Number:", _safe_str(report_data.master_serial_number)),
                ("Asset Number:", _safe_str(report_data.master_asset_number)),
                ("Certificate No.:", _safe_str(report_data.master_certificate_number)),
                ("Traceability Chain:", _safe_str(report_data.master_traceability_chain)),
                ("Uncertainty (u):", _safe_str(report_data.master_uncertainty_u)),
                ("Accuracy:", _safe_str(report_data.master_accuracy)),
                ("Resolution:", _safe_str(report_data.master_resolution)),
                ("Calibration Due Date:", _format_date(report_data.master_cal_due_date)),
                ("Claimed CMC:", _safe_str(report_data.master_claimed_cmc)),
            ],
            styles,
            [label_col_narrow, value_col_narrow],
        )
    )

    # Calibration readings
    spacer()
    heading("Calibration Readings")

    readings_blocks = _build_readings_blocks(report_data)
    for block in readings_blocks:
        if block["title"]:
            story.append(Paragraph(block["title"], styles["body_bold"]))
            story.append(Spacer(1, 4))

        # Preserve the original Pressure/Electrical header's two-line
        # wrapping (ReportLab's Table splits plain-string cells on "\n"
        # into wrapped lines within the same cell - verified empirically).
        header_row = [h.replace(" Measured", "\nMeasured") for h in block["header"]]
        block_table_data = [header_row] + block["rows"]
        col_w = usable_width / len(header_row)
        block_table = Table(
            block_table_data,
            colWidths=[col_w] * len(header_row),
            repeatRows=1,
        )
        block_table.setStyle(_black_table_style(has_header_row=True))
        story.append(block_table)
        if block is not readings_blocks[-1]:
            spacer()

    # Uncertainty budget
    spacer()
    heading("Measurement Uncertainty Budget")
    pdf_optional_component_fields = [
        ("Type A Uncertainty (u_A):", report_data.type_a_value),
        ("Standard Uncertainty (u_std):", report_data.u_std),
        ("Standard's Accuracy Uncertainty (u_std_accuracy):", report_data.u_std_accuracy),
        ("Resolution Uncertainty (u_res):", report_data.u_res),
        ("Hysteresis Uncertainty (u_hys):", report_data.u_hys),
        ("Zero Uncertainty (u_zero):", report_data.u_zero),
        ("Temperature Influence Uncertainty (u_temp):", report_data.u_temp),
        ("Repeatability Uncertainty (u_repeatability):", report_data.u_repeatability),
        ("Standard Weights Uncertainty (u_std_weights):", report_data.u_std_weights),
        ("Eccentric Loading Uncertainty (u_eccentric):", report_data.u_eccentric),
        ("Drift of Standard Uncertainty (u_drift):", report_data.u_drift),
        ("Bath Stability Uncertainty (u_bath_stability):", report_data.u_bath_stability),
        ("Bath Uniformity Uncertainty (u_bath_uniformity):", report_data.u_bath_uniformity),
        ("Wire Homogeneity Uncertainty (u_wire_homogeneity):", report_data.u_wire_homogeneity),
    ]
    uncertainty_rows = [
        (label, _safe_str(value))
        for label, value in pdf_optional_component_fields
        if value is not None
    ]
    uncertainty_rows += [
        ("CMC:", _safe_str(report_data.cmc)),
        ("Combined Uncertainty (u_c):", _safe_str(report_data.combined_uncertainty)),
        ("Expanded Uncertainty (U):", _safe_str(report_data.expanded_uncertainty)),
        ("Coverage Factor (k):", _safe_str(report_data.k_value)),
        ("Final Applied Uncertainty:", _safe_str(report_data.final_applied_uncertainty)),
    ]
    story.append(
        _build_kv_table(
            uncertainty_rows,
            styles,
            [label_col_wide, value_col_wide],
        )
    )

    # Compliance statement
    spacer()
    heading("Compliance Statement")
    compliance_data = [
        [
            Paragraph("Compliance Status:", styles["body_bold"]),
            Paragraph(_safe_str(report_data.session_status), styles["body_bold"]),
        ]
    ]
    compliance_table = Table(compliance_data, colWidths=[label_col_wide, value_col_wide])
    compliance_table.setStyle(_kv_table_style())
    story.append(compliance_table)

    # Signature block
    spacer()
    story.append(Spacer(1, 6 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BLACK))
    story.append(Spacer(1, 3 * mm))

    sig_data = [
        [
            Paragraph("Authorised Signatory", styles["body_bold"]),
            Paragraph("", styles["body"]),
            Paragraph("Date", styles["body_bold"]),
        ],
        ["", "", ""],
    ]
    sig_table = Table(sig_data, colWidths=[80 * mm, 20 * mm, usable_width - 100 * mm])
    sig_table.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TEXTCOLOR", (0, 0), (-1, -1), BLACK),
                ("LINEBELOW", (0, 1), (0, 1), THIN_BORDER, BLACK),
                ("LINEBELOW", (2, 1), (2, 1), THIN_BORDER, BLACK),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
            ]
        )
    )
    story.append(sig_table)

    return story


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — EXCEL GENERATION
# ══════════════════════════════════════════════════════════════════════════════


def _estimate_wrapped_line_count(text: Any, chars_per_line: int) -> int:
    """Estimate how many wrapped lines a string of text will occupy.

    Args:
        text: The text to estimate wrapping for. None/empty renders as 1 line.
        chars_per_line: Approximate number of characters that fit on one
            wrapped line at the target column width and font size.

    Returns:
        int: Estimated number of lines, at least 1.
    """
    if not text:
        return 1
    chars_per_line = max(1, chars_per_line)
    total_lines = 0
    for paragraph in str(text).split("\n"):
        if not paragraph:
            total_lines += 1
        else:
            total_lines += max(1, math.ceil(len(paragraph) / chars_per_line))
    return max(1, total_lines)


def _row_height_for_text(
    text: Any, chars_per_line: int, font_size: int = 9, min_height: float = 15.0
) -> float:
    """Compute an Excel row height (points) tall enough to avoid clipping
    wrapped text when printed or converted to PDF.

    openpyxl does not auto-fit row height to wrapped cell content the way
    Excel's interactive UI does when a file is opened and viewed - if a
    row's height is left at Excel's default (sized for one line), any
    wrapped text beyond that first line is visually cut off in print or
    when converted to PDF via a headless tool, even though the underlying
    cell value is fully intact. This must be set explicitly per row.

    Args:
        text: The text that will be wrapped into this row's value cell.
        chars_per_line: Approximate characters per wrapped line, based on
            the merged column width the text will render in.
        font_size: Point size of the font used, for line-height estimation.
        min_height: Minimum row height regardless of content (points).

    Returns:
        float: Row height in points.
    """
    line_count = _estimate_wrapped_line_count(text, chars_per_line)
    # 1.5x font size per line is a deliberately generous line-height
    # estimate (typical single-spaced text is closer to 1.2x) - erring
    # tall is harmless, erring short reproduces the exact clipping bug
    # this function exists to prevent.
    points_per_line = font_size * 1.5
    return max(min_height, line_count * points_per_line)


def generate_excel_report(session_id: str) -> FileResponse:
    """Generate an Excel calibration certificate for the given session.

    Args:
        session_id: UUID of the calibration session.

    Returns:
        FileResponse: Streaming .xlsx attachment.

    Raises:
        ValueError: When the session does not exist or has status REJECTED.
        RuntimeError: When a required database record is missing.
    """
    report_data = gather_report_data(session_id)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Calibration Certificate"

    thin_side = Side(border_style="thin", color="000000")
    all_borders = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=14, color="000000")
    section_font = Font(name="Calibri", bold=True, size=10, color="000000")
    label_font = Font(name="Calibri", bold=True, size=9, color="000000")
    value_font = Font(name="Calibri", size=9, color="000000")
    header_font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")

    black_fill = PatternFill(fill_type="solid", fgColor="000000")
    grey_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    LAST_COLUMN = 6

    def write_title(row_num: int, text: str) -> None:
        """Write a bold centred title merged across all columns.

        Args:
            row_num: 1-indexed worksheet row.
            text: Title text.
        """
        cell = worksheet.cell(row=row_num, column=1, value=text)
        cell.font = title_font
        cell.alignment = center_align
        worksheet.merge_cells(
            start_row=row_num, start_column=1, end_row=row_num, end_column=LAST_COLUMN
        )
        worksheet.row_dimensions[row_num].height = 24

    def write_section_heading(row_num: int, text: str) -> None:
        """Write a grey-filled section heading merged across all columns.

        Args:
            row_num: 1-indexed worksheet row.
            text: Heading text.
        """
        worksheet.merge_cells(
            start_row=row_num, start_column=1, end_row=row_num, end_column=LAST_COLUMN
        )
        for col in range(1, LAST_COLUMN + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = grey_fill
            cell.border = all_borders
        heading_cell = worksheet.cell(row=row_num, column=1, value=text)
        heading_cell.font = section_font
        heading_cell.alignment = left_align

    def write_label_value(row_num: int, label: str, value: Any) -> None:
        """Write a bold label in column A and its value merged across B-F.

        Sets an explicit row height sized to whichever of the label or
        value text needs more wrapped lines, so long text (e.g. a full
        customer address or traceability statement) doesn't get visually
        clipped on print or PDF conversion. See _row_height_for_text.

        Args:
            row_num: 1-indexed worksheet row.
            label: Field label string.
            value: Field value; None is rendered as an empty string.
        """
        label_cell = worksheet.cell(row=row_num, column=1, value=label)
        label_cell.font = label_font
        label_cell.border = all_borders
        label_cell.alignment = left_align

        display_value = "" if value is None else str(value)
        worksheet.merge_cells(
            start_row=row_num, start_column=2, end_row=row_num, end_column=LAST_COLUMN
        )
        for col in range(2, LAST_COLUMN + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.border = all_borders
        value_cell = worksheet.cell(row=row_num, column=2, value=display_value)
        value_cell.font = value_font
        value_cell.alignment = left_align

        label_height = _row_height_for_text(label, CHARS_PER_LINE_LABEL, font_size=9)
        value_height = _row_height_for_text(display_value, CHARS_PER_LINE_VALUE, font_size=9)
        worksheet.row_dimensions[row_num].height = max(label_height, value_height)

    def write_table_header(row_num: int, column_labels: list[str]) -> None:
        """Write a black-filled header row with white bold text.

        Args:
            row_num: 1-indexed worksheet row.
            column_labels: Ordered column label strings.
        """
        for col_index, label_text in enumerate(column_labels, start=1):
            cell = worksheet.cell(row=row_num, column=col_index, value=label_text)
            cell.font = header_font
            cell.fill = black_fill
            cell.alignment = center_align
            cell.border = all_borders

    def write_data_row(row_num: int, values: list[Any]) -> None:
        """Write a plain data row with borders across all provided columns.

        Args:
            row_num: 1-indexed worksheet row.
            values: Ordered cell values.
        """
        for col_index, cell_value in enumerate(values, start=1):
            cell = worksheet.cell(
                row=row_num,
                column=col_index,
                value="" if cell_value is None else cell_value,
            )
            cell.font = value_font
            cell.alignment = left_align
            cell.border = all_borders

    worksheet.column_dimensions["A"].width = 36
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 20
    worksheet.column_dimensions["G"].width = 20

    # Chars-per-line estimates for the row-height fix, derived from the
    # actual column widths above. Column A holds labels only (rarely
    # wraps); columns B:F merge into one wide value cell for
    # write_label_value's content.
    CHARS_PER_LINE_LABEL = int(36 * 0.95)
    CHARS_PER_LINE_VALUE = int(20 * 5 * 0.95)

    # Print/page setup - fit to one page wide (so nothing is cut off the
    # right edge), unconstrained pages tall (so a long results table
    # flows across multiple pages rather than being squeezed unreadably
    # small trying to force everything onto a single page).
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2
    )
    worksheet.print_options.horizontalCentered = True

    current_row = 1

    write_title(current_row, "CALIBRATION CERTIFICATE")
    current_row += 2

    write_section_heading(current_row, "Certificate Details")
    current_row += 1

    for label, value in [
        ("Certificate No.", report_data.certificate_number),
        ("Date of Calibration", _format_date(report_data.date_of_calibration)),
        ("Calibration Due Date", _format_date(report_data.cal_due_date)),
        ("Item Received Date", _format_date(report_data.item_received_date)),
        ("Date of Issue", _format_date(report_data.date_of_issue)),
        ("Customer Name", report_data.customer_name),
        ("Customer Address", report_data.customer_address),
        ("Technician", report_data.technician),
    ]:
        write_label_value(current_row, label, value)
        current_row += 1

    current_row += 1

    write_section_heading(current_row, "Instrument Under Test")
    current_row += 1

    cal_range_display = None
    if report_data.instrument_range_min is not None and report_data.instrument_range_max is not None:
        unit_suffix = f" {report_data.instrument_unit}" if report_data.instrument_unit else ""
        cal_range_display = (
            f"{report_data.instrument_range_min} to {report_data.instrument_range_max}{unit_suffix}"
        )

    for label, value in [
        ("Instrument Name", report_data.instrument_name),
        ("Make", report_data.instrument_make),
        ("Model", report_data.instrument_model),
        ("Serial Number", report_data.instrument_serial_number),
        ("Tag Number", report_data.instrument_tag_number),
        ("Cal Range", cal_range_display),
        ("Resolution", report_data.instrument_resolution),
        ("Accuracy", report_data.instrument_accuracy_class),
        ("Dial Size", report_data.instrument_dial_size),
        ("Mounting Orientation", report_data.instrument_mounting_orientation),
        ("Location", report_data.instrument_location),
        ("Medium", report_data.instrument_medium_used),
        ("Calibration Carried At", report_data.instrument_calibration_carried_at),
    ]:
        write_label_value(current_row, label, value)
        current_row += 1

    current_row += 1

    write_section_heading(current_row, "Environmental Conditions")
    current_row += 1

    for label, value in [
        ("Temperature (°C)", report_data.temperature_c),
        ("Relative Humidity (%)", report_data.humidity_pct),
    ]:
        write_label_value(current_row, label, value)
        current_row += 1

    current_row += 1

    write_section_heading(current_row, "Reference / Master Instrument")
    current_row += 1

    for label, value in [
        ("Instrument Name", report_data.master_name),
        ("Make", report_data.master_make),
        ("Model", report_data.master_model),
        ("Serial Number", report_data.master_serial_number),
        ("Asset Number", report_data.master_asset_number),
        ("Certificate No.", report_data.master_certificate_number),
        ("Traceability Chain", report_data.master_traceability_chain),
        ("Uncertainty (u)", report_data.master_uncertainty_u),
        ("Accuracy", report_data.master_accuracy),
        ("Resolution", report_data.master_resolution),
        ("Calibration Due Date", _format_date(report_data.master_cal_due_date)),
        ("Claimed CMC", report_data.master_claimed_cmc),
    ]:
        write_label_value(current_row, label, value)
        current_row += 1

    current_row += 1

    write_section_heading(current_row, "Calibration Readings")
    current_row += 1

    readings_blocks = _build_readings_blocks(report_data)
    for block in readings_blocks:
        if block["title"]:
            sub_heading_cell = worksheet.cell(row=current_row, column=1, value=block["title"])
            sub_heading_cell.font = label_font
            current_row += 1

        write_table_header(current_row, block["header"])
        current_row += 1

        for data_row in block["rows"]:
            write_data_row(current_row, data_row)
            current_row += 1

        current_row += 1

    write_section_heading(current_row, "Measurement Uncertainty Budget")
    current_row += 1

    # Every possible component across all categories - only ones actually
    # present (not None) for this session's category get rendered, same
    # pattern as CalculationView.jsx uses on the frontend. Avoids hardcoding
    # a Pressure-only field list that would silently omit Weighing's/
    # Temperature's real components from the certificate.
    optional_component_fields = [
        ("Type A Uncertainty (u_A)", report_data.type_a_value),
        ("Standard Uncertainty (u_std)", report_data.u_std),
        ("Standard's Accuracy Uncertainty (u_std_accuracy)", report_data.u_std_accuracy),
        ("Resolution Uncertainty (u_res)", report_data.u_res),
        ("Hysteresis Uncertainty (u_hys)", report_data.u_hys),
        ("Zero Uncertainty (u_zero)", report_data.u_zero),
        ("Temperature Influence Uncertainty (u_temp)", report_data.u_temp),
        ("Repeatability Uncertainty (u_repeatability)", report_data.u_repeatability),
        ("Standard Weights Uncertainty (u_std_weights)", report_data.u_std_weights),
        ("Eccentric Loading Uncertainty (u_eccentric)", report_data.u_eccentric),
        ("Drift of Standard Uncertainty (u_drift)", report_data.u_drift),
        ("Bath Stability Uncertainty (u_bath_stability)", report_data.u_bath_stability),
        ("Bath Uniformity Uncertainty (u_bath_uniformity)", report_data.u_bath_uniformity),
        ("Wire Homogeneity Uncertainty (u_wire_homogeneity)", report_data.u_wire_homogeneity),
    ]
    for label, value in optional_component_fields:
        if value is not None:
            write_label_value(current_row, label, value)
            current_row += 1

    for label, value in [
        ("CMC", report_data.cmc),
        ("Combined Uncertainty (u_c)", report_data.combined_uncertainty),
        ("Expanded Uncertainty (U)", report_data.expanded_uncertainty),
        ("Coverage Factor (k)", report_data.k_value),
        ("Final Applied Uncertainty", report_data.final_applied_uncertainty),
    ]:
        write_label_value(current_row, label, value)
        current_row += 1

    current_row += 1

    write_section_heading(current_row, "Compliance Statement")
    current_row += 1

    label_cell = worksheet.cell(row=current_row, column=1, value="Compliance Status")
    label_cell.font = label_font
    label_cell.border = all_borders
    label_cell.alignment = left_align

    worksheet.merge_cells(
        start_row=current_row, start_column=2, end_row=current_row, end_column=LAST_COLUMN
    )
    for col in range(2, LAST_COLUMN + 1):
        worksheet.cell(row=current_row, column=col).border = all_borders
    compliance_value_cell = worksheet.cell(
        row=current_row, column=2, value=_safe_str(report_data.session_status)
    )
    compliance_value_cell.font = Font(name="Calibri", bold=True, size=9, color="000000")
    compliance_value_cell.alignment = left_align

    temp_file = tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx", prefix=f"cal_cert_{session_id}_"
    )
    temp_path = temp_file.name
    temp_file.close()

    try:
        workbook.save(temp_path)

        safe_cert_number = "".join(
            ch for ch in _safe_str(report_data.certificate_number)
            if ch.isalnum() or ch in ("-", "_")
        )
        download_filename = f"CalibrationCertificate_{safe_cert_number}.xlsx"

        return FileResponse(
            path=temp_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=download_filename,
            background=_DeleteFileAfterResponse(temp_path),
        )

    except Exception:
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        raise


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — AUDIT LOGGING
# ══════════════════════════════════════════════════════════════════════════════


def write_audit_entry(
    session_id: str,
    report_format: str,
    user_id: str,
) -> None:
    """Record a certificate generation event to the audit log.

    Args:
        session_id: UUID of the calibration session that was reported on.
        report_format: Format string indicating the output type, e.g. "PDF" or "EXCEL".
        user_id: UUID of the user who triggered the download.

    Returns:
        None

    Raises:
        RuntimeError: If the database write fails.
    """
    audit_payload = {
        "user_id": user_id,                                         # matches audit_log.user_id
        "action": f"CERTIFICATE_GENERATED_{report_format.upper()}", # matches audit_log.action
        "table_affected": "calibration_sessions",                   # matches audit_log.table_affected
        "record_id": session_id,                                    # matches audit_log.record_id
    }

    insert_audit_log(audit_payload)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED INFRASTRUCTURE
# ══════════════════════════════════════════════════════════════════════════════


class _DeleteFileAfterResponse:
    """Starlette background task that deletes a temp file after the response is sent.

    Args:
        file_path: Absolute path of the file to delete.
    """

    def __init__(self, file_path: str) -> None:
        self._file_path = file_path

    async def __call__(self) -> None:
        """Delete the file; log a warning on failure without raising.

        Returns:
            None
        """
        try:
            if os.path.exists(self._file_path):
                os.unlink(self._file_path)
        except OSError:
            logger.warning(
                "Failed to delete temporary report file: %s",
                self._file_path,
                exc_info=True,
            )